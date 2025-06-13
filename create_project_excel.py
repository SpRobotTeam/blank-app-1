import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, BarChart, Reference
import os

def create_project_excel():
    """프로젝트 분석 엑셀 파일 생성 (고급 포맷팅 포함)"""
    
    # 데이터 정의
    project_overview = {
        '항목': ['프로젝트명', '총 프로젝트 금액', '기계 어셈블리', '제어전장', '설치/시운전', '후처리'],
        '값': ['GANTY-LODER 자동 용접 시스템', '980,181,870원', '640,181,870원', '90,000,000원', '170,000,000원', '80,000,000원']
    }
    
    main_analysis = {
        '구분': ['기계 어셈블리', '기계 어셈블리', '기계 어셈블리', '기계 어셈블리', '기계 어셈블리', '기계 어셈블리', '기계 어셈블리', '기계 어셈블리', '전장 시스템', '설치/시운전', '설치/시운전', '후처리', '총계'],
        '항목': ['SADDLE', 'CARRIAGE', 'Y-AXIS', 'Z-AXIS', 'X-AXIS', 'R-AXIS', '잡자재', '소계', '제어전장 구매품', '기계설치시운전', '제어설치시운전', '후처리 작업', '총계'],
        '총금액': [53070336, 15159090, 71698323, 30442500, 17198621, 452583000, 30000, 640181870, 90000000, 120000000, 50000000, 80000000, 980181870],
        '제작서비스': [4977336, 2164090, 24698323, 2442500, 9198621, 7583000, 0, 50063870, 0, 120000000, 50000000, 80000000, 300063870],
        '구매금액': [48093000, 12995000, 47000000, 28000000, 8000000, 445000000, 30000, 590118000, 90000000, 0, 0, 0, 680118000],
        '제작비율': ['9.4%', '14.3%', '34.4%', '8.0%', '53.5%', '1.7%', '0%', '7.8%', '0%', '100%', '100%', '100%', '30.6%']
    }
    
    expensive_items = {
        '품목': ['ABB 로봇 시스템', 'LINCOLN 용접기 R450', 'Y-AXIS GIRDER', '감속기 (YJD110)', 'APEX 감속기 (AFR140)', 'THK LM GUIDE', '한국이구스 에너지체인', 'THK 볼스크류'],
        '금액': [110000000, 34800000, 19500000, 2758000, 1795000, 1340720, 3400000, 2370000],
        '비고': ['R-AXIS 핵심 장비', '용접 핵심 장비', '제작 품목', 'SADDLE용, 6개', 'CARRIAGE용, 3개', 'Y-AXIS용, 6개', 'Z-AXIS용', 'Z-AXIS용']
    }
    
    manufacturing_items = {
        '품목': ['X-RAIL', 'X-BEAM', 'POST-BODY', 'Y-AXIS GIRDER', 'CARRIAGE-PLATE', 'Z-BEAM'],
        '금액': [3675000, 4241500, 851250, 19500000, 1497000, 1857750],
        '수량': [1, 1, 18, 1, 1, 1],
        '어셈블리': ['X-AXIS', 'X-AXIS', 'X-AXIS', 'Y-AXIS', 'CARRIAGE', 'Z-AXIS']
    }
    
    category_ratios = {
        '분야': ['기계 어셈블리', '제어전장', '설치/시운전', '후처리'],
        '금액': [640181870, 90000000, 170000000, 80000000],
        '비율': ['65.3%', '9.2%', '17.3%', '8.2%']
    }

    # 엑셀 파일 생성
    with pd.ExcelWriter('GANTY_LODER_프로젝트_BOM_분석.xlsx', engine='openpyxl') as writer:
        # 각 시트에 데이터 쓰기
        pd.DataFrame(project_overview).to_excel(writer, sheet_name='프로젝트 개요', index=False)
        pd.DataFrame(main_analysis).to_excel(writer, sheet_name='메인 분석표', index=False)
        pd.DataFrame(expensive_items).to_excel(writer, sheet_name='고가 구매품목', index=False)
        pd.DataFrame(manufacturing_items).to_excel(writer, sheet_name='주요 제작품목', index=False)
        pd.DataFrame(category_ratios).to_excel(writer, sheet_name='분야별 비율', index=False)
        
        # 워크북 객체 가져오기
        workbook = writer.book
        
        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 각 시트 스타일링
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # 헤더 스타일 적용
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # 모든 셀에 보더 적용
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    if cell.row > 1:  # 헤더가 아닌 경우
                        cell.alignment = Alignment(horizontal='right' if isinstance(cell.value, (int, float)) else 'left')
            
            # 컬럼 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    print("✅ 엑셀 파일이 성공적으로 생성되었습니다: GANTY_LODER_프로젝트_BOM_분석.xlsx")
    return "GANTY_LODER_프로젝트_BOM_분석.xlsx"

if __name__ == "__main__":
    create_project_excel()
