import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_ganty_loader_excel():
    """GANTY-LODER 프로젝트 분석 데이터를 엑셀 파일로 생성"""
    
    # 1. 프로젝트 개요 데이터
    project_overview = {
        '항목': ['프로젝트명', '총 프로젝트 금액', '제작/서비스 금액', '구매 금액', '제작 비율'],
        '내용': ['GANTY-LODER 자동 용접 시스템', '980,181,870원', '300,063,870원', '680,118,000원', '30.6%']
    }
    
    # 2. 기계 어셈블리 상세 데이터
    assembly_data = {
        '어셈블리': ['SADDLE', 'CARRIAGE', 'Y-AXIS', 'Z-AXIS', 'X-AXIS', 'R-AXIS', '잡자재'],
        '총 금액 (원)': [53070336, 15159090, 71698323, 30442500, 17198621, 452583000, 30000],
        '제작/서비스 (원)': [4977336, 2164090, 24698323, 2442500, 9198621, 7583000, 0],
        '구매 금액 (원)': [48093000, 12995000, 47000000, 28000000, 8000000, 445000000, 30000],
        '제작 비율 (%)': [9.4, 14.3, 34.4, 8.0, 53.5, 1.7, 0]
    }
    
    # 3. 전체 프로젝트 구성 데이터
    project_composition = {
        '구분': ['기계 어셈블리', '전장 시스템', '설치/시운전', '후처리', '총계'],
        '총 금액 (원)': [640181870, 90000000, 170000000, 80000000, 980181870],
        '제작/서비스 (원)': [50063870, 0, 170000000, 80000000, 300063870],
        '구매 금액 (원)': [590118000, 90000000, 0, 0, 680118000],
        '제작 비율 (%)': [7.8, 0, 100, 100, 30.6],
        '전체 비율 (%)': [65.3, 9.2, 17.3, 8.2, 100]
    }
    
    # 4. 고가 구매 품목 데이터
    expensive_items = {
        '품목명': ['ABB 로봇 시스템', 'LINCOLN 용접기 R450', 'Y-AXIS GIRDER', 'THK 가이드/볼스크류', '영진웜 감속기', 'APEX 감속기'],
        '금액 (원)': [110000000, 34800000, 19500000, 15000000, 8500000, 6200000],
        '구분': ['구매', '구매', '제작', '구매', '구매', '구매'],
        '비고': ['핵심 로봇', '용접 장비', '대형 구조물', '정밀 이송계', '감속기', '감속기']
    }
    
    # 5. 서비스 상세 내역
    service_details = {
        '서비스 구분': ['기계설치시운전', '제어설치시운전', '후처리 작업'],
        '금액 (원)': [120000000, 50000000, 80000000],
        '비율 (%)': [12.2, 5.1, 8.2],
        '주요 내용': [
            '현장 조립, 정밀 얼라인먼트, 기계적 동작 시험',
            '제어시스템 설치, 프로그래밍, 통신 설정',
            '도장/표면처리, 최종 검사, 문서 작성'
        ]
    }
    
    # Excel 파일 생성
    file_path = 'data/ganty_loader_project_analysis.xlsx'
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # 각 시트에 데이터 저장
        pd.DataFrame(project_overview).to_excel(writer, sheet_name='프로젝트 개요', index=False)
        pd.DataFrame(assembly_data).to_excel(writer, sheet_name='기계 어셈블리', index=False)
        pd.DataFrame(project_composition).to_excel(writer, sheet_name='전체 구성', index=False)
        pd.DataFrame(expensive_items).to_excel(writer, sheet_name='고가 품목', index=False)
        pd.DataFrame(service_details).to_excel(writer, sheet_name='서비스 내역', index=False)
    
    # 스타일링 적용
    apply_excel_styling(file_path)
    
    print(f"엑셀 파일이 생성되었습니다: {file_path}")
    return file_path

def apply_excel_styling(file_path):
    """엑셀 파일에 스타일링 적용"""
    wb = openpyxl.load_workbook(file_path)
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 헤더 스타일 적용
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # 데이터 셀에 테두리 및 정렬 적용
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if cell.column in [2, 3, 4]:  # 금액 컬럼들
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = center_alignment
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)

if __name__ == "__main__":
    create_ganty_loader_excel()
